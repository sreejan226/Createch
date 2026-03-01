"""
Bill of Quantities: Excel inventory input and BoQ Excel output.
Runs the optimizer with demo data + uploaded inventory (consumables); BoQ is
derived from the optimization result (time-aware purchase orders).
"""
from __future__ import annotations

from io import BytesIO
from typing import Dict, List

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from api.demo import Mode as DemoMode, generate_demo_input
from domain.models import PurchaseOrder
from optimizer.solver import compute_purchase_orders
from services.demo_to_domain import demo_to_optimization_input

# Fallback unit prices if not in optimization result
DEFAULT_PRICES: Dict[str, float] = {
    "ALU_DECK_1x1": 85.0,
    "ALU_DECK_0.5x1": 55.0,
    "ALU_WALL_1x2": 120.0,
    "ALU_WALL_0.5x2": 85.0,
    "ALU_BEAM_0.3x2": 95.0,
    "PROP": 40.0,
    "BEAM": 60.0,
    "TIE_ROD": 12.0,
    "CLAMP": 4.0,
}

router = APIRouter(prefix="/boq", tags=["bill-of-quantities"])


def _parse_inventory_excel(content: bytes) -> Dict[str, int]:
    """Parse Excel: first column = material, second column = quantity. Returns dict material -> quantity."""
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    inventory: Dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if not row or row[0] is None:
            continue
        material = str(row[0]).strip()
        if not material or material.upper() in ("MATERIAL", "ITEM", "MATERIAL NAME"):
            continue
        try:
            qty = int(float(row[1])) if row[1] is not None else 0
        except (TypeError, ValueError):
            qty = 0
        if material:
            inventory[material] = inventory.get(material, 0) + qty
    wb.close()
    return inventory


def _build_boq_excel(
    purchase_orders: List[PurchaseOrder],
    prices: Dict[str, float],
) -> bytes:
    """Build BoQ Excel from purchase_orders: Material | Quantity Required | Unit Price | Total Price, then total row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill of Quantities"

    headers = ["Material", "Quantity Required", "Unit Price", "Total Price", "Earliest bucket"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    total_amount = 0.0
    row = 2
    for po in sorted(purchase_orders, key=lambda p: p.item):
        unit_price = prices.get(po.item, 0.0)
        line_total = po.quantity * unit_price
        total_amount += line_total
        ws.cell(row=row, column=1, value=po.item)
        ws.cell(row=row, column=2, value=po.quantity)
        ws.cell(row=row, column=3, value=unit_price)
        ws.cell(row=row, column=4, value=round(line_total, 2))
        ws.cell(row=row, column=5, value=po.arrival_bucket)
        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="(No additional materials required)")
        row = 3

    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_amount, 2))
    ws.cell(row=row, column=4).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_boq_from_inventory_content(
    content: bytes,
    *,
    seed: int = 7,
    horizon_days: int = 90,
    n_sites: int = 3,
    n_panels: int = 200,
    n_tasks: int = 40,
    mode: DemoMode = "BALANCED",
) -> bytes:
    """
    Parse inventory Excel, build demo project data (using given demo params), then
    compute purchase orders (demand minus consumables) and output BoQ Excel.
    Changing seed/horizon_days/n_sites/n_panels/n_tasks/mode changes the demo and thus the BoQ.
    """
    consumables = _parse_inventory_excel(content)
    demo = generate_demo_input(
        seed=seed,
        horizon_days=horizon_days,
        n_sites=n_sites,
        n_panels=n_panels,
        n_tasks=n_tasks,
        mode=mode,
    )
    opt_input = demo_to_optimization_input(demo, consumables)
    purchase_orders = compute_purchase_orders(opt_input)
    prices = dict(opt_input.costs.purchase) if opt_input.costs.purchase else DEFAULT_PRICES
    return _build_boq_excel(purchase_orders, prices)


@router.get(
    "/template",
    summary="Download BoQ inventory Excel template",
    response_description="Excel template with columns: Material, Quantity",
)
def download_boq_template() -> Response:
    """
    Download a simple Excel template for inventory:

    - Column A: Material (e.g. ALU_DECK_1x1, ALU_WALL_1x2, BEAM, PROP, CLAMP)
    - Column B: Quantity

    All known materials used in the demo (ALU_DECK, ALU_WALL, ALU_BEAM, PROP, BEAM, TIE_ROD, CLAMP)
    are pre-populated as rows, but you can add/remove rows freely.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Template"

    headers = ["Material", "Quantity"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Dynamically add one row per known material so the template always matches DEFAULT_PRICES.
    row = 2
    for material in sorted(DEFAULT_PRICES.keys()):
        ws.cell(row=row, column=1, value=material)
        ws.cell(row=row, column=2, value=0)
        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_template.xlsx"},
    )


@router.post(
    "/from-inventory",
    summary="Generate BoQ from inventory Excel",
    response_description="Excel file: Bill of Quantities (materials to purchase, quantity, price, total)",
)
async def generate_boq_from_inventory(
    file: UploadFile = File(..., description="Excel file: first column = Material, second column = Quantity"),
    seed: int = 7,
    horizon_days: int = 90,
    n_sites: int = 3,
    n_panels: int = 200,
    n_tasks: int = 40,
    mode: DemoMode = "BALANCED",
):
    """
    Upload an Excel file with **current inventory**:
    - **Column 1**: Material name (e.g. ALU_DECK_1x1, PROP, CLAMP)
    - **Column 2**: Quantity

    **Demo parameters** (query/form): change these to change the project used for demand;
    different seed/sites/tasks produce different requirements and thus a different BoQ.
    - **seed**: random seed (e.g. 7, 42) — different seed = different tasks/quantities
    - **horizon_days**: project horizon (e.g. 30, 90)
    - **n_sites**: number of sites (e.g. 2, 5)
    - **n_panels**: number of panels in demo
    - **n_tasks**: number of tasks — more tasks = more demand
    - **mode**: LOW_COST | BALANCED | FAST (affects safety stock in overrides)

    The response is an Excel **Bill of Quantities** file (download).
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be an Excel file (.xlsx or .xls)")

    content = await file.read()
    if not content:
        raise HTTPException(400, "Uploaded file is empty")

    try:
        boq_bytes = build_boq_from_inventory_content(
            content,
            seed=seed,
            horizon_days=horizon_days,
            n_sites=n_sites,
            n_panels=n_panels,
            n_tasks=n_tasks,
            mode=mode,
        )
    except Exception as e:  # pragma: no cover - defensive
        raise HTTPException(400, f"Could not parse Excel or build BoQ: {e!s}") from e

    return Response(
        content=boq_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bill_of_quantities.xlsx"},
    )

